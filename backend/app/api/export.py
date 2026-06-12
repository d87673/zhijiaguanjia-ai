"""数据导出 API — CSV 和 Excel 格式"""
import io
import csv
from datetime import datetime, timezone
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from sqlalchemy.orm import selectinload

from app.core.database import get_db
from app.core.deps import get_current_user, get_current_company
from app.models.models import User, Order, OrderItem, Customer, Staff, Service, Payment

router = APIRouter(prefix="/export", tags=["Export"])

SUPPORTED_FORMATS = ["csv", "excel"]


def _format_datetime(dt):
    if dt is None:
        return ""
    return dt.isoformat()


async def _get_company_orders(company_id: str, db: AsyncSession):
    """获取公司所有订单（含关联数据）"""
    result = await db.execute(
        select(Order)
        .where(Order.company_id == company_id)
        .options(selectinload(Order.customer))
        .options(selectinload(Order.assigned_staff))
        .options(selectinload(Order.items).selectinload(OrderItem.service))
        .options(selectinload(Order.payment))
        .order_by(Order.created_at.desc())
    )
    return result.scalars().all()


async def _get_company_customers(company_id: str, db: AsyncSession):
    """获取公司所有客户（含订单计数）"""
    result = await db.execute(
        select(Customer)
        .where(Customer.company_id == company_id)
        .options(selectinload(Customer.orders))
        .order_by(Customer.created_at.desc())
    )
    return result.scalars().all()


def _build_csv(headers: list[str], rows: list[list[str]]) -> str:
    """构建 CSV 字符串"""
    output = io.StringIO(newline="")
    writer = csv.writer(output)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return output.getvalue()


def _build_excel(sheets: dict[str, tuple[list[str], list[list[str]]]]) -> bytes:
    """构建 Excel 文件（多 Sheet）"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="Excel 导出需要安装 openpyxl 库")

    wb = Workbook()
    # 删除默认 sheet
    wb.remove(wb.active)

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for idx, (sheet_name, (headers, rows)) in enumerate(sheets.items()):
        if idx == 0:
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(title=sheet_name)

        # 写表头
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # 写数据行
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        # 自适应列宽（近似）
        for col_idx in range(1, len(headers) + 1):
            max_width = len(headers[col_idx - 1]) * 2  # 中文字符大约占 2 单位
            for row in rows[:100]:  # 取前 100 行估算
                cell_val = str(row[col_idx - 1]) if col_idx - 1 < len(row) else ""
                max_width = max(max_width, len(cell_val) * 1.5)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_width + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ═══════════════════════════════════════════════════
# 订单导出
# ═══════════════════════════════════════════════════

@router.get("/orders")
async def export_orders(
    format: str = Query("csv", pattern="^(csv|excel)$"),
    company_id: str = Depends(get_current_company),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    orders = await _get_company_orders(company_id, db)

    headers = [
        "订单ID", "客户姓名", "客户电话", "员工姓名", "状态",
        "服务项目", "数量", "单价", "总金额",
        "支付方式", "支付状态", "实付金额", "支付时间",
        "预约时间", "服务地址", "备注", "创建时间", "完成时间",
    ]

    rows = []
    for o in orders:
        items_str = "、".join(
            f"{item.service.name if item.service else '服务'}×{item.quantity}"
            for item in (o.items or [])
        )
        total_qty = sum(item.quantity for item in (o.items or []))
        avg_price = float(o.total_amount) / total_qty if total_qty > 0 else 0

        rows.append([
            str(o.id)[:8],
            o.customer.name if o.customer else "",
            o.customer.phone if o.customer else "",
            o.assigned_staff.name if o.assigned_staff else "",
            o.status,
            items_str,
            str(total_qty),
            f"{avg_price:.2f}",
            f"{float(o.total_amount):.2f}",
            o.payment.method if o.payment else "",
            o.payment.status if o.payment else "未支付",
            f"{float(o.payment.amount):.2f}" if o.payment else "0.00",
            _format_datetime(o.payment.paid_at) if o.payment else "",
            _format_datetime(o.scheduled_at),
            o.address or "",
            o.notes or "",
            _format_datetime(o.created_at),
            _format_datetime(o.completed_at),
        ])

    if format == "csv":
        csv_content = _build_csv(headers, rows)
        return StreamingResponse(
            io.BytesIO(csv_content.encode("utf-8-sig")),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=orders_{datetime.now(timezone.utc).strftime('%Y%m%d')}.csv"},
        )
    else:
        excel_bytes = _build_excel({"订单明细": (headers, rows)})
        return StreamingResponse(
            io.BytesIO(excel_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=orders_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"},
        )


# ═══════════════════════════════════════════════════
# 客户导出
# ═══════════════════════════════════════════════════

@router.get("/customers")
async def export_customers(
    format: str = Query("csv", pattern="^(csv|excel)$"),
    company_id: str = Depends(get_current_company),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    customers = await _get_company_customers(company_id, db)

    headers = [
        "客户ID", "姓名", "电话", "邮箱", "地址",
        "标签", "订单数", "备注", "创建时间",
    ]

    rows = []
    for c in customers:
        rows.append([
            str(c.id)[:8],
            c.name,
            c.phone or "",
            c.email or "",
            c.address or "",
            "、".join(c.tags) if c.tags else "",
            str(len(c.orders) if c.orders else 0),
            c.notes or "",
            _format_datetime(c.created_at),
        ])

    if format == "csv":
        csv_content = _build_csv(headers, rows)
        return StreamingResponse(
            io.BytesIO(csv_content.encode("utf-8-sig")),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=customers_{datetime.now(timezone.utc).strftime('%Y%m%d')}.csv"},
        )
    else:
        excel_bytes = _build_excel({"客户信息": (headers, rows)})
        return StreamingResponse(
            io.BytesIO(excel_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=customers_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"},
        )


# ═══════════════════════════════════════════════════
# 综合导出（多 Sheet Excel）
# ═══════════════════════════════════════════════════

@router.get("/full")
async def export_full(
    company_id: str = Depends(get_current_company),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """导出全部数据为多 Sheet Excel"""

    # ── 订单 ──
    orders = await _get_company_orders(company_id, db)
    order_headers = [
        "订单ID", "客户姓名", "客户电话", "员工姓名", "状态",
        "服务项目", "总金额", "预约时间", "服务地址", "创建时间",
    ]
    order_rows = []
    for o in orders:
        items_str = "、".join(
            f"{item.service.name if item.service else '服务'}×{item.quantity}"
            for item in (o.items or [])
        )
        order_rows.append([
            str(o.id)[:8], o.customer.name if o.customer else "",
            o.customer.phone if o.customer else "",
            o.assigned_staff.name if o.assigned_staff else "",
            o.status, items_str, f"{float(o.total_amount):.2f}",
            _format_datetime(o.scheduled_at), o.address or "",
            _format_datetime(o.created_at),
        ])

    # ── 客户 ──
    customers = await _get_company_customers(company_id, db)
    cust_headers = ["客户ID", "姓名", "电话", "邮箱", "地址", "标签", "订单数", "创建时间"]
    cust_rows = []
    for c in customers:
        cust_rows.append([
            str(c.id)[:8], c.name, c.phone or "", c.email or "",
            c.address or "", "、".join(c.tags) if c.tags else "",
            str(len(c.orders) if c.orders else 0), _format_datetime(c.created_at),
        ])

    # ── 员工 ──
    staff_result = await db.execute(
        select(Staff).where(Staff.company_id == company_id).order_by(Staff.name)
    )
    all_staff = staff_result.scalars().all()
    staff_headers = ["员工ID", "姓名", "电话", "技能", "评分", "当前负载", "状态", "创建时间"]
    staff_rows = []
    for s in all_staff:
        staff_rows.append([
            str(s.id)[:8], s.name, s.phone or "",
            "、".join(s.skills) if s.skills else "",
            f"{s.rating:.1f}", str(s.current_load),
            "在岗" if s.is_active else "休息",
            _format_datetime(s.created_at),
        ])

    # ── 收入汇总 (月维度) ──
    rev_result = await db.execute(
        select(
            func.date_trunc("month", Payment.paid_at).label("month"),
            func.sum(Payment.amount).label("total"),
            func.count(Payment.id).label("count"),
        )
        .join(Order, Payment.order_id == Order.id)
        .where(Order.company_id == company_id, Payment.status == "paid")
        .group_by(func.date_trunc("month", Payment.paid_at))
        .order_by(func.date_trunc("month", Payment.paid_at).desc())
    )
    rev_rows_raw = rev_result.all()
    rev_headers = ["月份", "收入金额", "订单数"]
    rev_rows = [
        [r.month.strftime("%Y-%m") if r.month else "", f"{float(r.total):.2f}", str(r.count)]
        for r in rev_rows_raw
    ]

    excel_bytes = _build_excel({
        "订单明细": (order_headers, order_rows),
        "客户信息": (cust_headers, cust_rows),
        "员工信息": (staff_headers, staff_rows),
        "收入汇总": (rev_headers, rev_rows),
    })

    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=full_report_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"},
    )
